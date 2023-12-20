import requests
import pandas as pd
import logging
import time
import re
import json
import signal
import sys
import os
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# 18/11/23

interrupted = False

# Keywords to exclude from the list of games
excluded_keywords = ['hentai', 'sex', 'nazi', 'sexual', 'daddy', 'mommy', 'furry', 'college', 'fuck' , 'fucking', 'ass', 'dick' , 'penis', 'cock', 
                         'boob', 'tit', 'trump', 'obama', 'biden', 'anime', 'test', 'pack', 'dlc', 'soundtrack', 'demo']

# Configure logging
# To turn off info change logging level to logging.CRITICAL
# to see debugs set to debug (you wont need to)
# To stop saving file, remove the file handler
# If you want to stop terminal output, remove console handler
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler("steam_api.log"), # file handler 
                        logging.StreamHandler() # console handler
                    ])

def load_unwanted_games():
    try:
        with open('unwanted_games.json', 'r') as f:                
            unwanted_games = json.load(f)
            logging.info("Loaded list of games without reviews.")
            return unwanted_games
    except FileNotFoundError:
            unwanted_games = []
            logging.warning("No existing list of unwanted games found.")
            return unwanted_games


# This are game ids that are dlcs or have no reviews.
def save_unwanted_games(unwanted_games):
    try:
        filename = 'unwanted_games.json'
        # Check if file exists and is not empty
        if os.path.exists(filename) and os.path.getsize(filename) > 0:
            with open(filename, 'r+') as file:
                # Load existing data
                existing_data = json.load(file)
                # Update data
                updated_data = list(set(existing_data + unwanted_games))  # Remove duplicates
                # Rewind file to the beginning
                file.seek(0)
                # Dump updated data
                json.dump(updated_data, file)
                # Truncate file to new size
                file.truncate()
        else:
            # If file doesn't exist or is empty, write new data
            with open(filename, 'w') as file:
                json.dump(unwanted_games, file)

        logging.info(f"Saved unwanted games data to {filename}")

    except Exception as e:
        logging.critical(f"Error saving unwanted games data: {e}")  

# functions for saving and loading the appid list, so we can resume from where we left off
def save_appid_list(appid_data, filename='remaining_games.json'):
    try:
        with open(filename, 'w') as file:
            json.dump(appid_data, file)
        logging.info(f"Saved AppID data to {filename}")
    except Exception as e:
        logging.critical(f"Error saving AppID data: {e}")

def load_game_list(filename='remaining_games.json'):
    try:
        if os.path.exists(filename):
            with open(filename, 'r') as file:
                logging.info(f"Loading AppID data from {filename}")
                return json.load(file)
    except Exception as e:
        logging.critical(f"Error loading AppID data: {e}")
    return None

# Auto adjust column widths after writing to Excel since it's going to be Yuge
def auto_adjust_columns(filepath):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    wrap_columns = ['C', 'J'] # Word wrap for the description and review text columns

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[col_letter].width = length

        if col_letter in wrap_columns:
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True)

    workbook.save(filepath)
    workbook.close()
    logging.info("Column widths adjusted.")

# Function to write data to Excel
def write_data_to_excel(game_data, filename='steam_data.xlsx'):

    wrap_columns = ['C', 'J'] # Word wrap for the description and review text columns

    try:
        df_new_data = pd.DataFrame(game_data)

        if os.path.exists(filename):
            with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                if 'Steam Data' in pd.ExcelFile(filename).sheet_names:
                    df_existing = pd.read_excel(filename, sheet_name='Steam Data')
                    df_combined = pd.concat([df_existing, df_new_data], ignore_index=True)
                    df_combined.to_excel(writer, index=False, sheet_name='Steam Data')
                else:
                    df_new_data.to_excel(writer, index=False, sheet_name='Steam Data')

        auto_adjust_columns(filename, wrap_columns)

        logging.info(f"Data successfully saved to {filename}")

    except Exception as e:
        logging.critical(f"Failed to write data to Excel file: {e}")


# check for crt+c in terminal to stop the script and write the data to excel
def signal_handler(sig, frame):
    global interrupted
    interrupted = True
    logging.info("Ctrl+C detected. Preparing to save data.")     


def is_english(s):
    try:
        s.encode(encoding='utf-8').decode('ascii')
    except UnicodeDecodeError:
        return False
    else:
        return True

def contains_excluded_keywords(s, keywords):
    return any(keyword in s.lower() for keyword in keywords)

def get_steam_games():
    url = "https://api.steampowered.com/ISteamApps/GetAppList/v2/"
    response = requests.get(url)
    data = response.json()
    apps = data["applist"]["apps"]
    
    return [app for app in apps if app['name'].strip() 
            and is_english(app['name']) 
            and not contains_excluded_keywords(app['name'], excluded_keywords)]

def get_game_details(appid):
    url = f"http://store.steampowered.com/api/appdetails?appids={appid}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()

        logging.debug(f"API response for {appid}: {data}")  # Log the entire API response

        if str(appid) in data and data[str(appid)]['success']:
            data = data[str(appid)]['data']

            if 'dlc' in data: # I don't think this does anything but it's here
                logging.info(f"AppID {appid} is a DLC. Skipping.")
                return data, True  # flag for DLC
            
            elif any(category.get('id', 0) == 21 for category in data.get('categories', [])):  # ID 21 apparently indicates DLC
                logging.info(f"AppID {appid} is a DLC. Skipping.")
                return data, True # Flag for DLC

            return data, False 

        logging.warning(f"Response for AppID {appid} marked as unsuccessful.")
        return None, False

    except requests.HTTPError as http_err:
        if response.status_code == 429:
            retry_after = response.headers.get('Retry-After')
            logging.error(f"HTTP 429 Rate Limit hit. Headers: {response.headers}")
            if retry_after: # It never has this header but its supposed to so its here from my pain and suffering
                wait_time = int(retry_after)
                logging.info(f"Rate limit hit for game details API. Waiting for {wait_time} seconds before retrying.")
                time.sleep(wait_time)
                return get_game_details(appid)
            else:
                wait_time = 60
                logging.warning("Rate limit hit for game details API, but no Retry-After header found. Waiting for 60 seconds before retrying.")
                time.sleep(wait_time)
                return None, False
        else:
            logging.error(f"HTTP error {response.status_code} occurred while fetching details for AppID {appid}: {http_err}")
            logging.error(f"Response: {response.text}")
            return None, False

    except requests.RequestException as req_err:
        logging.warning(f"Request error occurred while fetching details for AppID {appid}: {req_err}")
        return None, False

    except ValueError as json_err:
        logging.critical(f"JSON decoding failed for details of AppID {appid}: {json_err}. Response: {response.text}")
        return None, False

def get_game_reviews_summary(appid):
    url = f"https://store.steampowered.com/appreviews/{appid}?json=1"
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()

    except requests.HTTPError as http_err:
        if response.status_code == 429:
            retry_after = response.headers.get('Retry-After')
            if retry_after: # It actually never has this header, but it's supposed to so I'm leaving it in for fun. 
                wait_time = int(retry_after)
                logging.info(f"Rate limit hit for reviews API. Waiting for {wait_time} seconds before retrying.")
                time.sleep(wait_time)
                return get_game_reviews_summary(appid)
            else:
                wait_time = 60
                logging.warning("Rate limit hit for reviews API, but no Retry-After header found. Waiting for 60 seconds before retrying.")
                time.sleep(wait_time)
                return get_game_reviews_summary(appid)
        else:
            logging.error(f"HTTP error {response.status_code} occurred while fetching reviews for AppID {appid}: {http_err}")
            logging.error(f"Response: {response.text}")

    except requests.RequestException as req_err:
        logging.warning(f"Request error occurred while fetching reviews for AppID {appid}: {req_err}")

    except ValueError as json_err:
        logging.critical(f"JSON decoding failed for reviews of AppID {appid}: {json_err}. Response: {response.text}")

    return None

def format_timestamp(timestamp):
    return datetime.utcfromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')

def main():
    game_data = []
    unwanted_games = []

    try:
        # Register the signal handler for Ctrl+C
        signal.signal(signal.SIGINT, signal_handler)

        load_unwanted_games()

        # Check to see if there is a list of remaning appids to process
        remaining_games = load_game_list()

        if remaining_games is not None:
            games = remaining_games
            games = [game for game in remaining_games # basically get rid of the games we don't want from the saved json and the word list
                 if game["appid"] not in unwanted_games # so we can update the word list and restart for future games
                 and not any(keyword.lower() in game["name"].lower() for keyword in excluded_keywords)]
            logging.info("Resuming from saved AppID list.")
            logging.info(f"Total games for processing: {len(games)}")
        else:
            games = get_steam_games() # otherwise get the list of games from the api
            games = [game for game in games if game["appid"] not in unwanted_games] # the function already checks for the word list
            logging.warning("No saved AppID list found. Fetching list of games from Steam API.")
            logging.info(f"Total games for processing: {len(games)}")

        # i is used for saving the remaining gameids for resuming the script
        for i in range(len(games)):
            game = games[i] 
            time.sleep(1.5) # rate limit is 200 requests per 5 minutes, so 1 request every 1.5 seconds should be fine

            # Check for Ctrl+C to stop the script and save data
            if interrupted:
                remaining_games = [{"appid": game["appid"], "name": game["name"]} for game in games[i:]]
                save_appid_list(remaining_games)
                write_data_to_excel(game_data)
                save_unwanted_games(unwanted_games)
                logging.info("Data and AppID list saved due to signal interrupt. Exiting the script.")
                sys.exit(0)
            
            appid = game["appid"]

            logging.info(f"Fetching details for {game['name']} (AppID: {appid})...")
            details, is_dlc = get_game_details(appid)

            # wanted seperation logging to make sure it was working
            if details is None:
                logging.debug(f"No valid details for {game['name']} (AppID: {appid}), skipping.")
                continue 

            if is_dlc:
                logging.debug(f"Skipping {game['name']} (AppID: {appid}) since it's a DLC.")
                unwanted_games.append(appid)
                continue

            reviews_summary = get_game_reviews_summary(appid)

            if reviews_summary is None: 
                logging.info(f"No valid reviews summary for {game['name']} (AppID: {appid}), skipping.")
                continue


            if details:
                name = details.get('name', 'No Name Available')
                genres = ", ".join([genre['description'] for genre in details['genres']]) if 'genres' in details else 'N/A'
                short_description = details.get('short_description', 'N/A')  # Extract the game description

                if 'query_summary' in reviews_summary and reviews_summary['query_summary']['total_reviews'] > 0:
                    qs = reviews_summary['query_summary']
                    positive_percentage = (qs['total_positive'] / qs['total_reviews']) * 100 if qs['total_reviews'] > 0 else 0
                    review = reviews_summary['reviews'][0] if reviews_summary['reviews'] else None
                    try:
                        review_text = review['review'] if review else 'N/A'
                        # Check if review text is in English, replace line breaks
                        if is_english(review_text):
                            review_text = review_text.replace('\n', ' ').replace('\r', '')
                        else:
                            review_text = "N/A"
                    except Exception as e:
                        logging.warning(f"Error processing review text for {name} (AppID: {appid}): {e}")
                        review_text = 'Error processing review text'

                    # qs is query summary 
                    game_data.append({
                        'Name': name,
                        'AppID': appid,
                        'Short Description': short_description,
                        'Review Score': qs['review_score'],
                        'Review Score Description': qs['review_score_desc'],
                        'Total Positive': qs['total_positive'],
                        'Total Negative': qs['total_negative'],
                        'Total Reviews': qs['total_reviews'],
                        'Positive Review Percentage': positive_percentage,
                        'Review Text (First Review)': review_text,
                        'Playtime Forever (First Review)': review['author'].get('playtime_forever', 'N/A') if review else 'N/A',
                        'Playtime Last Two Weeks (First Review)': review['author'].get('playtime_last_two_weeks', 'N/A') if review else 'N/A',
                        'Playtime at Review (First Review)': review['author'].get('playtime_at_review', 'N/A') if review else 'N/A',
                        'Timestamp Created (First Review)': format_timestamp(review['timestamp_created']) if review else 'N/A',
                        'Timestamp Updated (First Review)': format_timestamp(review['timestamp_updated']) if review else 'N/A',
                        'Voted Up (First Review)': review.get('voted_up', 'N/A') if review else 'N/A',
                        'Votes Up (First Review)': review.get('votes_up', 'N/A') if review else 'N/A',
                        'Genres': genres
                    })
                else:
                    logging.info(f"No reviews for {name} (AppID: {appid}), skipping.")
                    unwanted_games.append(appid)
            else:
                logging.warning(f"Game details fetch unsuccessful for AppID: {appid}.")

    except Exception as e:
        logging.critical(f"An unexpected error occurred: {e}")
        logging.critical("Full traceback:")
        logging.critical(traceback.format_exc()) 
        logging.info("Saving the current data before exiting.")

        remaining_games = [{"appid": game["appid"], "name": game["name"]} for game in games[i:]]
        save_appid_list(remaining_games)
        write_data_to_excel(game_data)
        save_unwanted_games(unwanted_games)        

    # Save the updated games_without_reviews list
    # Save the remaining appids list
    else:
        write_data_to_excel(game_data)
        save_unwanted_games(unwanted_games)

if __name__ == "__main__":
    main()
